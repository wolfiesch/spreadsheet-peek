"""Generate a wide-table stress-test workbook.

This sample exists to stress-test the token-cost claims at the opposite
end of the shape spectrum from `sample-financials.xlsx`. Wide tables
(29 columns) amplify box-drawing overhead because every row adds that
many column separators, so the per-row token cost of the default
`xleak -n` output balloons relative to the `--export text` path.

The synthetic data is a trailing-12-month operations dashboard: one
department per row, one metric per month (actual + plan), plus YoY
comparisons. Numbers are fabricated but structurally plausible.

Regenerate with:
    uv run --with openpyxl python examples/generate_wide_table.py
"""
from pathlib import Path
from openpyxl import Workbook

OUT_PATH = Path(__file__).parent / "wide-table.xlsx"

MONTHS = [
    "Apr 2023", "May 2023", "Jun 2023", "Jul 2023", "Aug 2023", "Sep 2023",
    "Oct 2023", "Nov 2023", "Dec 2023", "Jan 2024", "Feb 2024", "Mar 2024",
]

# Header row: Department + 12 months of actuals + 12 months of plan + 4 summary cols
HEADERS = (
    ["Department"]
    + [f"{m} Actual" for m in MONTHS]
    + [f"{m} Plan" for m in MONTHS]
    + ["TTM Actual", "TTM Plan", "Var %", "YoY %"]
)

DEPARTMENTS = [
    "Sales - Enterprise", "Sales - Mid-Market", "Sales - SMB",
    "Marketing - Growth", "Marketing - Brand", "Marketing - Events",
    "Product - Core", "Product - Platform", "Product - AI",
    "Engineering - Backend", "Engineering - Frontend", "Engineering - Infra",
    "Engineering - Data", "Engineering - Security", "Customer Success",
    "Support - L1", "Support - L2", "Professional Services",
    "Finance", "Legal", "People Ops", "IT", "Facilities", "Executive",
]


def deterministic_row(dept: str, seed_index: int) -> list:
    """Produce repeatable fake monthly numbers without importing random.

    A simple linear-congruential-ish formula keeps the output stable
    across regenerations, which matters because this file is committed.
    """
    base = 50_000 + seed_index * 7_500
    actuals = []
    plans = []
    for i, _ in enumerate(MONTHS):
        seasonal = 1.0 + 0.15 * ((i % 4) - 1.5) / 1.5  # Q1 dip, Q4 bump
        growth = 1.0 + 0.008 * i  # ~10% over 12 months
        actual = int(base * seasonal * growth)
        plan = int(base * seasonal * 1.10)  # plan is always +10%
        actuals.append(actual)
        plans.append(plan)

    ttm_actual = sum(actuals)
    ttm_plan = sum(plans)
    var_pct = (ttm_actual - ttm_plan) / ttm_plan * 100
    yoy_pct = 8.0 + (seed_index % 7) * 2.5  # 8-23% range

    return (
        [dept]
        + actuals
        + plans
        + [ttm_actual, ttm_plan, f"{var_pct:.1f}%", f"{yoy_pct:.1f}%"]
    )


def main() -> None:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    ws = wb.create_sheet("Dept Operations")

    ws.append(HEADERS)
    for idx, dept in enumerate(DEPARTMENTS):
        ws.append(deterministic_row(dept, idx))

    # Narrow columns so the file opens cleanly in Excel, but xleak's -w 30
    # default will still truncate - which is exactly the point.
    ws.column_dimensions["A"].width = 26
    for col_idx in range(2, len(HEADERS) + 1):
        col_letter = ws.cell(row=1, column=col_idx).coordinate.rstrip("0123456789")
        ws.column_dimensions[col_letter].width = 12

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} ({len(HEADERS)} columns x {len(DEPARTMENTS) + 1} rows)")


if __name__ == "__main__":
    main()
