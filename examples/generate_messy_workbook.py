"""Generate a messy real-world-style workbook for benchmark stress tests.

The existing samples are intentionally clean rectangles. This fixture is meant
to look more like a workbook an agent receives from a finance or operations
team: title rows, notes, blank separators, long memo text, dates, currencies,
percentages, mixed departments, and multiple sheets.

Regenerate with:
    uv run --with openpyxl python examples/generate_messy_workbook.py
"""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_PATH = Path(__file__).parent / "messy-ops-export.xlsx"
DATE_FORMAT = "yyyy-mm-dd"
CURRENCY_FORMAT = '$#,##0;[Red]($#,##0)'
PERCENT_FORMAT = "0.0%"

VENDORS = [
    "Northwind Data Services, LLC",
    "Acme Cloud Infrastructure",
    "Globex Fulfillment Partners",
    "Initech Support Desk",
    "Umbrella Security Review",
    "Stark Industrial Design",
    "Wayne Facilities Group",
    "Hooli Analytics Platform",
]

DEPARTMENTS = [
    "Finance",
    "Sales - Enterprise",
    "Sales - SMB",
    "Engineering - Platform",
    "Engineering - Data",
    "Customer Success",
    "Operations",
    "People Ops",
]

MEMOS = [
    "Monthly platform invoice; includes usage overage and tax true-up",
    "Renewal booked before procurement approval - needs follow-up",
    "Backdated credit memo applied to prior period",
    "Multi-line note from AP:\nconfirm whether this belongs in COGS",
    "Imported from card feed with merchant descriptor mismatch",
    "Manual reclass from department owner after close",
    "PO missing; keep in accrual review until approved",
    "Contains comma, pipe | and quote \" characters for parser sanity",
]


def spend_rows(count: int = 240) -> list[list[object]]:
    rows: list[list[object]] = []
    start = date(2024, 1, 2)

    for idx in range(count):
        base = 850 + ((idx * 791) % 42_000)
        if idx % 11 == 0:
            amount = -round(base * 0.18, 2)
            status = "Credit / reversal"
        else:
            amount = round(base * (1 + (idx % 5) * 0.07), 2)
            status = "Needs review" if idx % 13 == 0 else "Approved"

        confidence = 0.55 + ((idx * 17) % 45) / 100
        rows.append(
            [
                f"TXN-2024-{idx + 1:04d}",
                start + timedelta(days=idx % 91),
                VENDORS[idx % len(VENDORS)],
                DEPARTMENTS[(idx * 3) % len(DEPARTMENTS)],
                MEMOS[(idx * 5) % len(MEMOS)],
                "USD",
                amount,
                f"Cost center {(idx % 9) + 1}",
                status,
                confidence,
                f"owner-{(idx % 12) + 1}@example.com",
                "Imported" if idx % 4 else "Manual",
            ]
        )

    return rows


def write_spend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Spend Export")
    ws.append(["Q1 Vendor Spend Export - messy source workbook"])
    ws.append(["Generated for benchmark only", date(2024, 4, 8), "Do not treat as real company data"])
    ws.append(["Notes", "Rows include credits, long memos, multiline cells, punctuation, blanks, and review flags"])
    ws.append([])

    headers = [
        "Transaction ID",
        "Txn Date",
        "Vendor / Merchant Name",
        "Department",
        "Memo / Description",
        "Currency",
        "Amount",
        "Cost Center",
        "Review Status",
        "Confidence",
        "Owner",
        "Source",
    ]
    ws.append(headers)
    for row in spend_rows():
        ws.append(row)

    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A5"].font = Font(bold=True)
    for cell in ws[5]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")
        cell.alignment = Alignment(wrap_text=True)

    widths = {
        "A": 18,
        "B": 12,
        "C": 30,
        "D": 24,
        "E": 52,
        "F": 10,
        "G": 14,
        "H": 16,
        "I": 18,
        "J": 12,
        "K": 24,
        "L": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(6, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = DATE_FORMAT
        ws.cell(row=row, column=7).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=10).number_format = PERCENT_FORMAT
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:L{ws.max_row}"


def write_rollup_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Review Rollup")
    ws.append(["Department", "Transactions", "Gross Spend", "Credits", "Net Spend", "Needs Review %"])
    for idx, dept in enumerate(DEPARTMENTS):
        transactions = 24 + idx * 3
        gross = 125_000 + idx * 18_500
        credits = 4_500 + idx * 1_300
        net = gross - credits
        review_pct = 0.08 + idx * 0.015
        ws.append([dept, transactions, gross, credits, net, review_pct])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    for row in range(2, ws.max_row + 1):
        for col in (3, 4, 5):
            ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=6).number_format = PERCENT_FORMAT


def write_notes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Reviewer Notes")
    ws.append(["Topic", "Note"])
    ws.append(["Purpose", "Synthetic workbook for measuring spreadsheet preview token costs."])
    ws.append(["Messy traits", "Title rows, blank rows, long wrapped memos, mixed number formats, punctuation, and multiple sheets."])
    ws.append(["Scope", "The first sheet is intentionally source-like; the rollup sheet is intentionally report-like."])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 96
    for cell in ws[1]:
        cell.font = Font(bold=True)


def main() -> None:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    write_spend_sheet(wb)
    write_rollup_sheet(wb)
    write_notes_sheet(wb)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
