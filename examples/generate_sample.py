"""Generate a realistic multi-sheet financial spreadsheet for demo purposes.

This script creates examples/sample-financials.xlsx - a simplified monthly P&L,
Balance Sheet, and Revenue Breakdown. The data is entirely synthetic but
structured like a real small-business financial package.

Regenerate with:
    python examples/generate_sample.py
"""
from datetime import date
from pathlib import Path

from openpyxl import Workbook

OUT_PATH = Path(__file__).parent / "sample-financials.xlsx"

DATE_FORMAT = "yyyy-mm-dd"

# ---- P&L -------------------------------------------------------------------
PNL_HEADERS = ["Account", "Jan 2024", "Feb 2024", "Mar 2024", "Q1 Total", "Q1 2023", "YoY %"]
PNL_ROWS = [
    ["Revenue"],
    ["  Product Sales", 420_000, 445_000, 512_000, 1_377_000, 1_050_000, "31.1%"],
    ["  Services", 180_000, 195_000, 208_000, 583_000, 480_000, "21.5%"],
    ["  Subscriptions", 92_000, 97_500, 103_800, 293_300, 210_000, "39.7%"],
    ["Total Revenue", 692_000, 737_500, 823_800, 2_253_300, 1_740_000, "29.5%"],
    [""],
    ["Cost of Goods Sold", 277_000, 295_000, 329_500, 901_500, 732_000, "23.2%"],
    ["Gross Profit", 415_000, 442_500, 494_300, 1_351_800, 1_008_000, "34.1%"],
    ["Gross Margin %", "60.0%", "60.0%", "60.0%", "60.0%", "57.9%", ""],
    [""],
    ["Operating Expenses"],
    ["  Salaries & Benefits", 145_000, 148_000, 152_000, 445_000, 380_000, "17.1%"],
    ["  Rent", 18_000, 18_000, 18_000, 54_000, 48_000, "12.5%"],
    ["  Marketing", 32_000, 41_000, 38_500, 111_500, 78_000, "43.0%"],
    ["  Software & Tools", 12_400, 13_200, 13_800, 39_400, 31_000, "27.1%"],
    ["  Professional Fees", 8_500, 6_200, 14_000, 28_700, 22_000, "30.5%"],
    ["Total OpEx", 215_900, 226_400, 236_300, 678_600, 559_000, "21.4%"],
    [""],
    ["Operating Income", 199_100, 216_100, 258_000, 673_200, 449_000, "49.9%"],
    ["Operating Margin %", "28.8%", "29.3%", "31.3%", "29.9%", "25.8%", ""],
]

# ---- Balance Sheet ---------------------------------------------------------
BS_HEADERS = ["Account", date(2024, 3, 31), date(2023, 12, 31), "Change $", "Change %"]
BS_ROWS = [
    ["ASSETS"],
    ["Current Assets"],
    ["  Cash & Equivalents", 1_842_500, 1_120_000, 722_500, "64.5%"],
    ["  Accounts Receivable", 412_800, 385_000, 27_800, "7.2%"],
    ["  Inventory", 298_000, 265_000, 33_000, "12.5%"],
    ["  Prepaid Expenses", 45_200, 38_500, 6_700, "17.4%"],
    ["Total Current Assets", 2_598_500, 1_808_500, 790_000, "43.7%"],
    [""],
    ["Fixed Assets"],
    ["  Equipment (net)", 185_000, 210_000, -25_000, "-11.9%"],
    ["  Leasehold Improvements", 92_000, 98_500, -6_500, "-6.6%"],
    ["Total Fixed Assets", 277_000, 308_500, -31_500, "-10.2%"],
    [""],
    ["TOTAL ASSETS", 2_875_500, 2_117_000, 758_500, "35.8%"],
    [""],
    ["LIABILITIES & EQUITY"],
    ["  Accounts Payable", 185_400, 198_000, -12_600, "-6.4%"],
    ["  Accrued Expenses", 92_500, 78_000, 14_500, "18.6%"],
    ["  Deferred Revenue", 215_000, 178_000, 37_000, "20.8%"],
    ["Total Liabilities", 492_900, 454_000, 38_900, "8.6%"],
    [""],
    ["  Common Stock", 100_000, 100_000, 0, "0.0%"],
    ["  Retained Earnings", 2_282_600, 1_563_000, 719_600, "46.0%"],
    ["Total Equity", 2_382_600, 1_663_000, 719_600, "43.3%"],
    [""],
    ["TOTAL LIABILITIES & EQUITY", 2_875_500, 2_117_000, 758_500, "35.8%"],
]

# ---- Revenue Breakdown -----------------------------------------------------
REV_HEADERS = ["Customer", "Segment", "Jan", "Feb", "Mar", "Q1 Total", "% of Total"]
REV_ROWS = [
    ["Acme Corp", "Enterprise", 85_000, 92_000, 98_000, 275_000, "12.2%"],
    ["Globex Industries", "Enterprise", 72_000, 74_500, 81_000, 227_500, "10.1%"],
    ["Initech", "Mid-Market", 45_000, 48_000, 52_000, 145_000, "6.4%"],
    ["Umbrella LLC", "Mid-Market", 38_500, 41_000, 44_500, 124_000, "5.5%"],
    ["Stark Industries", "Enterprise", 62_000, 68_000, 75_000, 205_000, "9.1%"],
    ["Wayne Enterprises", "Enterprise", 58_000, 61_000, 64_000, 183_000, "8.1%"],
    ["Hooli Inc", "Mid-Market", 42_000, 43_500, 46_000, 131_500, "5.8%"],
    ["Pied Piper", "SMB", 18_500, 22_000, 25_500, 66_000, "2.9%"],
    ["Dunder Mifflin", "SMB", 15_000, 16_500, 18_000, 49_500, "2.2%"],
    ["Other (72 customers)", "Mixed", 256_000, 270_000, 320_000, 846_000, "37.6%"],
    [""],
    ["TOTAL", "", 692_000, 737_500, 823_800, 2_253_300, "100.0%"],
]


def write_sheet(ws, headers, rows, *, date_header_cols=()):
    ws.append(headers)
    for row in rows:
        ws.append(row)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.value == "":
                continue
            if isinstance(cell.value, date):
                cell.number_format = DATE_FORMAT

    for col_idx in date_header_cols:
        ws.cell(row=1, column=col_idx).number_format = DATE_FORMAT

    # Rough column widths for readability in Excel
    widths = {1: 28, 2: 14, 3: 14, 4: 14, 5: 14, 6: 14, 7: 12}
    for col_idx, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width


def main():
    wb = Workbook()

    # openpyxl creates a default sheet; rename it as our first sheet
    wb.remove(wb.active) if wb.active is not None else None
    pnl = wb.create_sheet("P&L")
    write_sheet(pnl, PNL_HEADERS, PNL_ROWS)

    bs = wb.create_sheet("Balance Sheet")
    write_sheet(bs, BS_HEADERS, BS_ROWS, date_header_cols=(2, 3))

    rev = wb.create_sheet("Revenue Breakdown")
    write_sheet(rev, REV_HEADERS, REV_ROWS)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
