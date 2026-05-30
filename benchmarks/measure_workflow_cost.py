"""Measure what an agent session pays to inspect a messy workbook.

The normal token benchmark measures output size. This script includes the
command or generated code too, which is closer to what lands in an agent
transcript when the agent either runs `wolfxl` or writes throwaway Python.

Run with:
    uv run --with tiktoken --with openpyxl python benchmarks/measure_workflow_cost.py
"""
from __future__ import annotations

import argparse
import shlex
import subprocess
from dataclasses import dataclass
from pathlib import Path

import tiktoken

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_FILE = ROOT / "examples" / "messy-ops-export.xlsx"
ENC = tiktoken.get_encoding("cl100k_base")


@dataclass(frozen=True)
class Row:
    path: str
    rows: int
    setup_text: str
    output: str
    note: str


def token_count(text: str) -> int:
    return len(ENC.encode(text))


def first_lines(output: str, limit: int) -> str:
    return "".join(output.splitlines(keepends=True)[:limit])


def run(cmd: list[str]) -> str:
    result = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if result.returncode != 0:
        details = result.stderr.strip() or result.stdout.strip() or "no output"
        raise RuntimeError(
            f"Command failed with exit code {result.returncode}: {shlex.join(cmd)}\n{details}"
        )
    return result.stdout


def openpyxl_tuple_dump(path: Path, rows: int) -> tuple[str, str]:
    import openpyxl  # type: ignore[import-not-found]

    code = "\n".join(
        [
            "import openpyxl",
            f"wb = openpyxl.load_workbook({str(path)!r}, data_only=True)",
            "ws = wb.active",
            f"for row in ws.iter_rows(max_row={rows}, values_only=True):",
            "    print(row)",
            "wb.close()",
            "",
        ]
    )
    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        worksheet = workbook.active
        if worksheet is None:
            raise RuntimeError("workbook has no active worksheet")
        output = "".join(
            f"{row}\n"
            for row in worksheet.iter_rows(max_row=rows, values_only=True)
        )
    finally:
        workbook.close()
    return code, output


def workflow_rows(path: Path, row_limits: tuple[int, ...]) -> list[Row]:
    result: list[Row] = []
    for rows in row_limits:
        box_cmd = ["wolfxl", "peek", str(path), "-n", str(rows)]
        result.append(
            Row(
                "wolfxl box preview",
                rows,
                shlex.join(box_cmd),
                run(box_cmd),
                "readable table with borders",
            )
        )

        text_cmd = ["wolfxl", "peek", str(path), "--export", "text"]
        result.append(
            Row(
                "wolfxl text preview",
                rows,
                f"{shlex.join(text_cmd)} | sed -n '1,{rows + 1}p'",
                first_lines(run(text_cmd), rows + 1),
                "same current sheet, bounded for repeat previews",
            )
        )

        code, output = openpyxl_tuple_dump(path, rows)
        result.append(
            Row(
                "generated openpyxl tuple dump",
                rows,
                code,
                output,
                "raw Python tuples; includes generated code cost",
            )
        )
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Measure agent workflow token costs.")
    parser.add_argument("--file", type=Path, default=DEFAULT_FILE, help="Workbook to measure.")
    parser.add_argument(
        "--rows",
        default="5,15,50",
        help="Comma-separated preview row counts. Default: 5,15,50.",
    )
    args = parser.parse_args()
    path = args.file.resolve()
    if not path.exists():
        raise SystemExit(f"file not found: {path}")
    row_limits = tuple(int(value.strip()) for value in args.rows.split(",") if value.strip())

    rows = workflow_rows(path, row_limits)

    print(f"Agent workflow token cost for `{path}`")
    print()
    print("| Path | Rows | Command/code tokens | Output tokens | Total tokens | Notes |")
    print("|------|-----:|--------------------:|--------------:|-------------:|-------|")
    totals: dict[tuple[str, int], int] = {}
    for row in rows:
        setup_tokens = token_count(row.setup_text)
        output_tokens = token_count(row.output)
        total = setup_tokens + output_tokens
        totals[(row.path, row.rows)] = total
        print(
            f"| {row.path} | {row.rows} | {setup_tokens:,} | {output_tokens:,} | "
            f"{total:,} | {row.note} |"
        )

    print()
    print("## Pairwise savings")
    for rows_count in row_limits:
        box_total = totals[("wolfxl box preview", rows_count)]
        text_total = totals[("wolfxl text preview", rows_count)]
        tuple_total = totals[("generated openpyxl tuple dump", rows_count)]
        print(
            f"- **{rows_count} rows**: text preview saves **{box_total - text_total:,} tokens** "
            f"vs box preview and **{tuple_total - text_total:,} tokens** vs generated openpyxl tuple dump."
        )


if __name__ == "__main__":
    main()
