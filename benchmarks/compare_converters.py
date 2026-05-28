"""Compare local spreadsheet-to-context paths on one file.

This is a local comparison harness, not a universal performance claim. It always
measures Spreadsheet Peek's local paths. If optional tools are installed, it
measures those too; missing tools are reported as skipped rather than treated as
failures.

Run with:
    uv run --with tiktoken --with openpyxl python benchmarks/compare_converters.py
    uv run --with tiktoken --with openpyxl --with 'markitdown[xlsx]' --with agent-xlsx --with click python benchmarks/compare_converters.py
"""
from __future__ import annotations

import argparse
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import tiktoken

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_FILE = ROOT / "examples" / "sample-financials.xlsx"
ENC = tiktoken.get_encoding("cl100k_base")


@dataclass(frozen=True)
class Mode:
    name: str
    scope: str
    command: list[str]
    required: bool = False
    max_lines: int | None = None
    runner: Callable[[Path], tuple[str, str, str]] | None = None


def modes(path: Path) -> list[Mode]:
    return [
        Mode(
            "Spreadsheet Peek text export",
            "current sheet values",
            ["wolfxl", "peek", str(path), "--export", "text"],
            required=True,
        ),
        Mode(
            "Spreadsheet Peek map",
            "workbook structure",
            ["wolfxl", "map", str(path), "--format", "text"],
            required=True,
        ),
        Mode(
            "Spreadsheet Peek agent summary",
            "bounded workbook summary",
            ["wolfxl", "agent", str(path), "--max-tokens", "800"],
            required=True,
        ),
        Mode(
            "Vanilla openpyxl tuple dump",
            "first 5 physical rows",
            [],
            runner=lambda candidate: run_openpyxl_tuple_dump(candidate, rows=5),
        ),
        Mode(
            "Vanilla openpyxl tuple dump",
            "first 15 physical rows",
            [],
            runner=lambda candidate: run_openpyxl_tuple_dump(candidate, rows=15),
        ),
        Mode("MarkItDown", "whole-file Markdown conversion", ["markitdown", str(path)]),
        Mode("agent-xlsx probe", "workbook structure", ["agent-xlsx", "probe", str(path)]),
    ]


def first_lines(output: str, limit: int | None) -> str:
    if limit is None:
        return output
    return "".join(output.splitlines(keepends=True)[:limit])


def run_command(command: list[str], *, max_lines: int | None = None) -> tuple[str, str, str]:
    if shutil.which(command[0]) is None:
        return "skipped", "", f"{command[0]} is not installed"
    try:
        result = subprocess.run(command, capture_output=True, text=True, check=False)
    except OSError as error:
        return "error", "", str(error)
    if result.returncode != 0:
        details = result.stderr.strip() or result.stdout.strip() or "no output"
        return "error", "", details.splitlines()[0]
    return "ok", first_lines(result.stdout, max_lines), "ok"


def run_openpyxl_tuple_dump(path: Path, *, rows: int) -> tuple[str, str, str]:
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError:
        return "skipped", "", "openpyxl is not installed"

    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook.active
    if worksheet is None:
        return "error", "", "workbook has no active worksheet"
    output = "".join(
        f"{tuple(row)}\n"
        for row in worksheet.iter_rows(max_row=rows, values_only=True)
    )
    code = "\n".join(
        [
            "import openpyxl",
            f"wb = openpyxl.load_workbook({str(path)!r}, data_only=True)",
            "ws = wb.active",
            f"for row in ws.iter_rows(max_row={rows}, values_only=True):",
            "    print(row)",
            "",
        ]
    )
    return "ok", output, code


def token_count(text: str) -> int:
    return len(ENC.encode(text))


def main() -> None:
    parser = argparse.ArgumentParser(description="Compare local converter output sizes.")
    parser.add_argument("--file", type=Path, default=DEFAULT_FILE, help="Workbook or table file to compare.")
    args = parser.parse_args()
    path = args.file.resolve()
    if not path.exists():
        raise SystemExit(f"file not found: {path}")

    rows = []
    for mode in modes(path):
        if mode.runner is not None:
            status, output, setup_text = mode.runner(path)
            command = setup_text
        else:
            status, output, notes = run_command(mode.command, max_lines=mode.max_lines)
            command = " ".join(mode.command)
            setup_text = command
        if status != "ok":
            if mode.required:
                command_text = " ".join(mode.command) if mode.command else mode.name
                raise SystemExit(f"required command failed: {command_text}\n{notes if mode.runner is None else setup_text}")
            rows.append(
                {
                    "tool": mode.name,
                    "scope": mode.scope,
                    "command": " ".join(mode.command) if mode.command else mode.name,
                    "status": status,
                    "output_tokens": "",
                    "command_tokens": "",
                    "total_tokens": "",
                    "bytes": "",
                    "notes": notes if mode.runner is None else setup_text,
                }
            )
            continue
        command_tokens = token_count(setup_text)
        output_tokens = token_count(output)
        rows.append(
            {
                "tool": mode.name,
                "scope": mode.scope,
                "command": command if mode.command else "generated openpyxl snippet",
                "status": "ok",
                "output_tokens": f"{output_tokens:,}",
                "command_tokens": f"{command_tokens:,}",
                "total_tokens": f"{output_tokens + command_tokens:,}",
                "bytes": f"{len(output.encode('utf-8')):,}",
                "notes": f"{len(output.splitlines()):,} lines",
            }
        )

    print(f"Local comparison harness for `{path}`")
    print()
    print("| Tool | Scope | Command/code | Status | Output tokens | Command/code tokens | Total tokens | Bytes | Notes |")
    print("|------|-------|--------------|--------|--------------:|--------------------:|-------------:|------:|-------|")
    for row in rows:
        print(
            f"| {row['tool']} | {row['scope']} | `{row['command']}` | {row['status']} | "
            f"{row['output_tokens']} | {row['command_tokens']} | {row['total_tokens']} | "
            f"{row['bytes']} | {row['notes']} |"
        )


if __name__ == "__main__":
    main()
