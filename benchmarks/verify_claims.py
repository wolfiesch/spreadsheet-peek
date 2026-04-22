"""Verify non-token claims made by spreadsheet-peek docs.

The token benchmark proves the cost ratios. This script proves the behavioral
claims that can drift independently: direct format support, date/number-format
rendering, pipe hygiene, and the budgeted agent surface.

Run with:
    uv run --with openpyxl python benchmarks/verify_claims.py
"""
from __future__ import annotations

import re
import subprocess
import tempfile
import zipfile
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).parent.parent
EXAMPLES = ROOT / "examples"
FINANCIALS = EXAMPLES / "sample-financials.xlsx"
MESSY_CSV = EXAMPLES / "messy.csv"
SAMPLE_XLS = EXAMPLES / "sample-minimal.xls"
SAMPLE_XLSB = EXAMPLES / "sample-date.xlsb"
SAMPLE_ODS = EXAMPLES / "sample-minimal.ods"
XLSX_WORKBOOK_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
)
XLSM_WORKBOOK_CONTENT_TYPE = (
    "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
)


def run(cmd: list[str], *, check: bool = True) -> subprocess.CompletedProcess[str]:
    result = subprocess.run(
        cmd,
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    if check and result.returncode != 0:
        details = result.stderr.strip() or result.stdout.strip() or "no output"
        raise AssertionError(
            f"Command failed with exit code {result.returncode}: {' '.join(cmd)}\n{details}"
        )
    return result


def assert_contains(haystack: str, needle: str, context: str) -> None:
    if needle not in haystack:
        raise AssertionError(f"{context}: expected to find {needle!r}\n{haystack}")


def assert_not_contains(haystack: str, needle: str, context: str) -> None:
    if needle in haystack:
        raise AssertionError(f"{context}: did not expect to find {needle!r}\n{haystack}")


def write_macro_enabled_workbook(source: Path, target: Path) -> None:
    """Write a minimal macro-enabled OOXML package from the sample workbook."""
    with zipfile.ZipFile(source, "r") as src, zipfile.ZipFile(
        target, "w", zipfile.ZIP_DEFLATED
    ) as dst:
        for entry in src.infolist():
            data = src.read(entry.filename)
            if entry.filename == "[Content_Types].xml":
                text = data.decode("utf-8")
                text = text.replace(
                    XLSX_WORKBOOK_CONTENT_TYPE,
                    XLSM_WORKBOOK_CONTENT_TYPE,
                )
                if XLSM_WORKBOOK_CONTENT_TYPE not in text:
                    raise AssertionError("failed to mark workbook package as xlsm")
                data = text.encode("utf-8")
            dst.writestr(entry, data)


def check_xlsx_preview() -> None:
    out = run(["wolfxl", "peek", str(FINANCIALS), "-n", "2"]).stdout
    assert_contains(out, "wolfxl peek - Excel preview", "xlsx preview banner")
    assert_contains(out, "Sheet: P&L", "xlsx first sheet")


def check_xlsm_preview() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        xlsm = Path(tmp) / "sample-financials.xlsm"
        write_macro_enabled_workbook(FINANCIALS, xlsm)
        with zipfile.ZipFile(xlsm) as package:
            content_types = package.read("[Content_Types].xml").decode("utf-8")
        assert_contains(
            content_types,
            XLSM_WORKBOOK_CONTENT_TYPE,
            "xlsm package content type",
        )
        out = run(["wolfxl", "peek", str(xlsm), "-n", "2"]).stdout
    assert_contains(out, "wolfxl peek - Excel preview", "xlsm preview banner")
    assert_contains(out, "Sheet: P&L", "xlsm first sheet")


def check_direct_legacy_format_previews() -> None:
    cases = [
        (SAMPLE_XLS, "P&L", "Revenue", "xls preview"),
        (SAMPLE_XLSB, "Sheet1", "2021-01-01", "xlsb preview"),
        (SAMPLE_ODS, "P&L", "Revenue", "ods preview"),
    ]
    for path, sheet_name, marker, context in cases:
        out = run(["wolfxl", "peek", str(path), "-n", "2"]).stdout
        assert_contains(out, "wolfxl peek - Excel preview", f"{context} banner")
        assert_contains(out, f"Sheet: {sheet_name}", f"{context} sheet")
        assert_contains(out, marker, f"{context} marker")


def check_direct_delimited_previews() -> None:
    csv_out = run(["wolfxl", "peek", str(MESSY_CSV), "-n", "2"]).stdout
    assert_contains(csv_out, "Sheet: messy", "csv preview sheet")
    assert_contains(csv_out, "Acme, Inc.", "csv quoted comma")

    with tempfile.TemporaryDirectory() as tmp:
        tsv = Path(tmp) / "sample.tsv"
        tsv.write_text("name\tamount\nAlice\t10\nBob\t20\n", encoding="utf-8")
        tsv_out = run(["wolfxl", "peek", str(tsv), "-n", "1"]).stdout
        assert_contains(tsv_out, "Sheet: sample", "tsv preview sheet")
        assert_contains(tsv_out, "Alice", "tsv preview row")

        txt = Path(tmp) / "sample.txt"
        txt.write_text("name,amount\nAlice,10\nBob,20\n", encoding="utf-8")
        txt_out = run(["wolfxl", "peek", str(txt), "-n", "1"]).stdout
        assert_contains(txt_out, "Sheet: sample", "txt preview sheet")
        assert_contains(txt_out, "Alice", "txt preview row")


def check_date_rendering() -> None:
    out = run(
        [
            "wolfxl",
            "peek",
            str(FINANCIALS),
            "--sheet",
            "Balance Sheet",
            "--export",
            "text",
        ]
    ).stdout
    first_line = out.splitlines()[0]
    assert_contains(first_line, "2024-03-31", "Balance Sheet date header")
    assert_contains(first_line, "2023-12-31", "Balance Sheet date header")


def check_number_format_rendering() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        workbook = Path(tmp) / "formatted-values.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Formats"
        ws.append(["Metric", "Value"])
        ws.append(["Revenue", 1234.5])
        ws.append(["Margin", 0.125])
        ws["B2"].number_format = "$#,##0.00"
        ws["B3"].number_format = "0.0%"
        wb.save(workbook)

        out = run(["wolfxl", "peek", str(workbook), "-n", "2"]).stdout

    assert_contains(out, "$1,234.50", "currency number format")
    assert_contains(out, "12.5%", "percentage number format")


def check_sed_pipe_hygiene() -> None:
    result = run(
        [
            "bash",
            "-o",
            "pipefail",
            "-c",
            'wolfxl peek "$1" --export text | sed -n "1,6p"',
            "--",
            str(FINANCIALS),
        ]
    )
    combined = f"{result.stdout}\n{result.stderr}"
    assert_contains(result.stdout, "Account\tJan 2024", "sed-limited text preview")
    assert_not_contains(combined.lower(), "broken pipe", "sed-limited text preview")


def check_agent_budget_surface() -> None:
    out = run(["wolfxl", "agent", str(FINANCIALS), "--max-tokens", "800"]).stdout
    assert_contains(out, "WORKBOOK:", "agent briefing")
    match = re.search(r"# wolfxl agent: (?P<used>\d+)/800 tokens", out)
    if not match:
        raise AssertionError(f"agent output did not include token footer\n{out}")
    used = int(match.group("used"))
    if used > 800:
        raise AssertionError(f"agent output exceeded budget: {used}/800")


def check_wolfxl_version_floor() -> None:
    version = run(["wolfxl", "--version"]).stdout.strip()
    match = re.match(r"wolfxl (?P<major>\d+)\.(?P<minor>\d+)\.", version)
    if not match:
        raise AssertionError(
            f"could not parse wolfxl version: {version}. "
            "Update docs, WOLFXL_CLI_VERSION, and this script together."
        )
    major = int(match.group("major"))
    minor = int(match.group("minor"))
    if (major, minor) < (0, 8):
        raise AssertionError(
            f"claim verification requires wolfxl >= 0.8.0 for direct multi-format reads; got {version}. "
            "Update docs, WOLFXL_CLI_VERSION, and this script together."
        )


def main() -> None:
    checks = [
        check_wolfxl_version_floor,
        check_xlsx_preview,
        check_xlsm_preview,
        check_direct_legacy_format_previews,
        check_direct_delimited_previews,
        check_date_rendering,
        check_number_format_rendering,
        check_sed_pipe_hygiene,
        check_agent_budget_surface,
    ]

    for check in checks:
        check()
        print(f"ok - {check.__name__}")


if __name__ == "__main__":
    main()
